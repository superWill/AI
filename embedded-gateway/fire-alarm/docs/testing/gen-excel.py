#!/usr/bin/env python3
"""从 gb4717-gb16806-test-cases.md 生成测试执行表 Excel。
用法：/usr/bin/python3 gen-excel.py   （Homebrew Python 3.14 的 expat 损坏，须用系统 Python）
输出：test-cases-execution.xlsx（含 测试结果/是否通过下拉/测试人/日期/备注 执行列）
"""
import re, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule

HERE = os.path.dirname(os.path.abspath(__file__))
SRC = os.path.join(HERE, "gb4717-gb16806-test-cases.md")
OUT = os.path.join(HERE, "test-cases-template.xlsx")  # 只输出空白模板；test-cases-execution.xlsx 是实测工作表，严禁脚本覆盖

GROUP = {"FA":"火灾报警","LC":"报警控制(联动输出)","FT":"故障报警","SH":"屏蔽","SV":"监管报警",
         "ST":"自检","DP":"显示与查询","CK":"检查功能","PW":"电源","WD":"看门狗",
         "CM":"通信","DR":"存储单元(黑匣子)","OP":"操作级别","GE":"气体灭火控制器",
         "EX":"结构与外观","UI":"人机菜单(产品级)","TT":"型式试验"}

def clean(s):
    s = s.replace("**", "").strip()
    return s

rows = []
seen = set()
for ln in open(SRC, encoding="utf-8").read().split("\n"):
    m = re.match(r'^\|\s*\*\*([A-Z]{2}-\d{2}(?:\.\d+)?)\*\*\s*\|(.*)\|\s*$', ln)
    if not m: continue
    cid = m.group(1)
    if cid in seen: continue          # 只取正文首次出现（附录矩阵里的引用不算）
    cells = [c.strip() for c in m.group(2).split("|")]
    name = steps = verdict = basis = prio = ""
    main = ""
    note = ""
    if len(cells) == 5:               # 原子版：级|依据|主文本|判定|备注
        prio, basis, main, verdict, note = cells
    elif len(cells) == 4:               # 标准表：级|依据|名称+步骤|判定
        prio, basis, main, verdict = cells
    elif len(cells) == 3:             # OP：级|名称+步骤|判定
        prio, main, verdict = cells
        basis = "5.4.13 表1 / 5.4.14"
    elif len(cells) == 2:             # EX / TT
        if cid.startswith("EX"):
            main, basis = cells       # 名称+判定 | 依据
            prio, verdict = "P1", ""
        else:                          # TT：名称+报警侧 | 气灭侧
            main, verdict = cells[0], "气灭控制器条件：" + clean(cells[1])
            prio, basis = "P2", "见 TT 表"
    else:
        continue
    nm = re.match(r'\*\*(.+?)\*\*——(.*)$', main, re.S)
    if nm:
        name, steps = nm.group(1), nm.group(2)
    else:
        name, steps = main, "（见父案步骤）" if '.' in cid else main
        if '.' not in cid: name = ""
    if cid.startswith("EX") and not verdict:
        verdict, steps = steps, "（外观/设计检查项）"
    seen.add(cid)
    rows.append([cid, GROUP.get(cid[:2], ""), clean(name), clean(prio),
                 clean(basis), clean(steps), clean(verdict), clean(note)])

wb = Workbook()
ws = wb.active
ws.title = "测试执行表"
headers = ["序号","用例号","分组","名称","级","依据","步骤要点","判定准则(预期)",
           "测试结果","是否通过","测试人","日期","备注"]
widths  = [6, 9, 14, 22, 10, 16, 55, 55, 30, 10, 8, 11, 20]
hfill = PatternFill("solid", fgColor="1F3864")
hfont = Font(bold=True, color="FFFFFF", size=11)
thin = Border(*[Side(style="thin", color="BFBFBF")]*4)
wrap = Alignment(vertical="top", wrap_text=True)

for i, (h, w) in enumerate(zip(headers, widths), 1):
    c = ws.cell(1, i, h); c.fill = hfill; c.font = hfont
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.column_dimensions[c.column_letter].width = w

for r, row in enumerate(rows, 2):
    ws.cell(r, 1, r-1)
    for i, v in enumerate(row[:7], 2):
        ws.cell(r, i, v)
    if len(row) > 7 and row[7]:
        ws.cell(r, 13, row[7])
    for i in range(1, len(headers)+1):
        c = ws.cell(r, i); c.border = thin; c.alignment = wrap

n = len(rows) + 1
dv = DataValidation(type="list", formula1='"通过,不通过,不适用,待测,暂不测试"', allow_blank=True)
ws.add_data_validation(dv); dv.add(f"J2:J{n}")
ws.conditional_formatting.add(f"J2:J{n}",
    CellIsRule(operator="equal", formula=['"通过"'], fill=PatternFill("solid", fgColor="C6EFCE")))
ws.conditional_formatting.add(f"J2:J{n}",
    CellIsRule(operator="equal", formula=['"不通过"'], fill=PatternFill("solid", fgColor="FFC7CE")))
ws.freeze_panes = "C2"
ws.auto_filter.ref = f"A1:M{n}"

# 汇总页
s2 = wb.create_sheet("统计")
s2["A1"], s2["B1"] = "指标", "值"
s2["A2"], s2["B2"] = "用例总数", len(rows)
s2["A3"] = "已测(是否通过非空)"
s2["B3"] = f'=COUNTA(测试执行表!J2:J{n})-COUNTIF(测试执行表!J2:J{n},"待测")'
s2["A4"], s2["B4"] = "通过", f'=COUNTIF(测试执行表!J2:J{n},"通过")'
s2["A5"], s2["B5"] = "不通过", f'=COUNTIF(测试执行表!J2:J{n},"不通过")'
s2["A6"], s2["B6"] = "不适用", f'=COUNTIF(测试执行表!J2:J{n},"不适用")'
s2["A7"], s2["B7"] = "通过率(通过/已测)", f'=IF(B3=0,"-",B4/B3)'
for r in range(1, 8):
    s2.cell(r, 1).font = Font(bold=(r == 1))
s2.column_dimensions["A"].width = 22; s2.column_dimensions["B"].width = 14

wb.save(OUT)
print(f"rows: {len(rows)} -> {OUT}")
missing = [k for k in GROUP if not any(r[0].startswith(k) for r in rows)]
print("缺组:", missing or "无")
